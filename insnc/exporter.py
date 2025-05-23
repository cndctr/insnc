# insnc/exporter.py

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def format_date(raw):
    return datetime.strptime(raw, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")

def export_operations_to_excel(items, filename=None):
    processed_ids = set()
    final_records = []

    # Group transfers
    for i, item in enumerate(items):
        id_ = item["id"]
        date = item["date"]
        amount = item["amount"]["amount"]

        if "TRANSFER" in item["icon"]["iconUrl"] and id_ not in processed_ids:
            for j, candidate in enumerate(items):
                if i == j or candidate["id"] in processed_ids:
                    continue
                if (
                    candidate["date"] == date and
                    candidate["amount"]["amount"] == -amount and
                    "TRANSFER" in candidate["icon"]["iconUrl"]
                ):
                    record = {
                        "date": format_date(date),
                        "title": candidate["description"] if candidate["amount"]["amount"] > 0 else item["description"],
                        "description": candidate["description"] if candidate["amount"]["amount"] < 0 else item["description"],
                        "Expense": abs(amount),
                        "Income": abs(amount),
                        "postfix": item["amount"]["postfix"],
                        "type": "Перевод"
                    }
                    final_records.append(record)
                    processed_ids.update({id_, candidate["id"]})
                    break

    # Group currency exchanges
    for i, item in enumerate(items):
        id_ = item["id"]
        date = item["date"]

        if "CURRENCY_EXCHANGE" in item["icon"]["iconUrl"] and id_ not in processed_ids:
            for j, candidate in enumerate(items):
                if i == j or candidate["id"] in processed_ids:
                    continue

                if (
                    candidate["date"] == date
                    and "CURRENCY_EXCHANGE" in candidate["icon"]["iconUrl"]
                ):
                    # Find the pair: one with positive amount, one with negative operationAmount
                    if item["amount"]["amount"] > 0 and "operationAmount" in item:
                        income_item = item
                        expense_item = candidate
                    elif candidate["amount"]["amount"] > 0 and "operationAmount" in candidate:
                        income_item = candidate
                        expense_item = item
                    else:
                        continue  # Not a valid pair

                    if "operationAmount" not in income_item or income_item["operationAmount"]["amount"] >= 0:
                        continue  # Must have negative operationAmount to reflect expense

                    record = {
                        "date": format_date(date),
                        "title": income_item["description"],
                        "description": expense_item["description"],
                        "Expense": abs(income_item["operationAmount"]["amount"]),
                        "Income": abs(income_item["amount"]["amount"]),
                        "postfix": "",
                        "type": "Конвертация"
                    }
                    final_records.append(record)
                    processed_ids.update({income_item["id"], expense_item["id"]})
                    break

    # Regular income/expense
    for item in items:
        id_ = item["id"]
        if id_ in processed_ids:
            continue

        amt_raw = item["amount"]["amount"]
        amt = item["operationAmount"]["amount"] if "operationAmount" in item and amt_raw < 0 else amt_raw
        postfix = item["operationAmount"]["postfix"] if "operationAmount" in item and amt_raw < 0 else item["amount"]["postfix"]

        record = {
            "date": format_date(item["date"]),
            "title": item["title"],
            "description": item["description"],
            "Expense": abs(amt) if amt < 0 else None,
            "Income": abs(amt) if amt > 0 else None,
            "postfix": postfix,
            "type": "Расход" if amt < 0 else "Приход"
        }

        final_records.append(record)

    # Sort by date
    final_records.sort(key=lambda r: r["date"])

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Operations"

    # Write header
    headers = ["date", "title", "description", "Expense", "Income", "postfix", "type"]
    ws.append(headers)

    # Write rows
    for record in final_records:
        row = [record.get(h) for h in headers]
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    if not filename:
        filename = "operation_history.xlsx"

    wb.save(filename)
    print(f"[✓] Exported to '{filename}'")
